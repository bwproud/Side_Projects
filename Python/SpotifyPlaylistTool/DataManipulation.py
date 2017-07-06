# -*- coding: utf-8 -*-
'''
Created on Jul 15, 2016
TODO: input dialog GUI, pretty print,
'''
import openpyxl
from blist import sortedset

#gets playlist
def getPlaylist(s):
    wb = openpyxl.load_workbook(s)
    sheet = wb.active
    return sheet

#gets simple spotify spreadsheet with just names and songs
def getArtistsv2():
    li=[]
    sheet= getPlaylist('Artists.xlsx')
    for i in range(1, sheet.max_row+1):
        r = 'A'+str(i)
        li.append((sheet[r].value))
    return li

#gets a sorted set of playlistCuration of simple spotify playlist       
def getSortedSet(sortS=False, songs=False):
    li=getArtistsv2()
    ss =[]
    s=li[0][36:37]
    if not songs:
        for i in li: 
            r = i[i.find(s)+2:]
            ss.append(r)
    else:
        for i in li:
            r = i[:i.find(s)]
            ss.append(r.strip())          
    return sortedset(ss) if sortS else ss 

#removes features
def truncateArtists(ss, sortedSet=True):
    li=[]
    for i in range(len(ss)):
        s=ss[i].strip()
        index = s.find(",")
        index2 = s.find(" feat")
        if index!=-1: s=s[:index]
        if index2!=-1: s=s[:index2]
        li.append(s)
    return sortedset(li) if sortedSet else li

def truncatePlusArtist(ss):
    li=[]
    for i in range(len(ss)):
        s=ss[i].strip()
        index = s.find(" - ")
        if index!=-1: s=s[:index]
        li.append(s)
    return li

#gets information from more complex spotify spreadsheet with
#songs, artist name, and album name
def getInfo(s):
    li=[]
    sheet = getPlaylist('playlistInfo.xlsx')
    for i in range(2, sheet.max_row+1):
        r = s+str(i)
        if s == 'B':
            li.append("%s"%sheet[r].value.strip())
        else: li.append("%s"%sheet[r].value)    
    return li

#gets songs
def getSongs(plusArtist = True):
    if not plusArtist:
        return getInfo('A')
    else:
        di = createDict(getSongs(plusArtist=False), getArtists())
        li = getInfo('A')
        songs =[]
        for i in range(len(li)):
            songs.append("%s - %s"%(li[i],di[li[i]]))
        return songs

#gets playlistCuration(includes option to make list a sorted set)
def getArtists(ss = False):    
    return sortedset(getInfo('B')) if ss else getInfo('B')
          
#gets albums(includes option to make list a sorted set)
def getAlbums(plusArtist=True, ss = False):
    sort = ss
    if not plusArtist:
        return sortedset(getInfo('C')) if ss else getInfo('C')
    else:
        di = createDict(getAlbums(plusArtist=False), truncateArtists(getArtists(), sortedSet=False))
        li = sortedset(getInfo('C')) if ss else getInfo('C')
        albums =[]
        for i in range(len(li)):
            albums.append("%s - %s"%(li[i],di[li[i]]))
        return albums
    
#creates a dictionary(HashMap) for use in adding artist names
#to spreadsheet to coorespond with album names    
def createDict(li, li2):
    di ={}
    for i in range(len(li)):
        di[li[i]]=li2[i]
    return di 
 
#create new spreadsheet with sorted set of both playlistCuration and albums        
def createArtistSpreadsheet():
    di=createDict(getAlbums(plusArtist=False), getArtists())
    wb = openpyxl.Workbook()
    dest_filename = 'SpotifyPlaylist.xlsx'
    ws1=wb.active
    ws1.title = "Artists"
    li=truncateArtists(getArtistsv2(), sortedSet=True) 
    ws1['A1']="Artists"
    for col in range(2,len(li)+2):
        ws1.cell(coordinate='A'+str(col), value=li[col-2])
    ws2 = wb.create_sheet(title="Albums")    
    li = getAlbums(ss=True)
    ws2['A1']="Albums"
    for col in range(2,len(li)+2):
        artist= di[li[col-2]]
        ws2.cell(coordinate='A'+str(col), value=li[col-2])
        ws2.cell(coordinate='B'+str(col), value=artist)  
    #wb.save(filename=dest_filename) 
    return "done"   

#finds duplicate songs in playlist
def findDupSongs(songs, artists):
    li = []
    dup =[]
    s= songs
    a= artists
    for i in range(len(s)):
        li.append("Artist: %s    Song: %s"%(a[i], s[i]))
    ss = list(sortedset(li))
    for i in range(len(s)):
        if li[i] in ss: ss.remove(li[i])
        else: dup.append(li[i])
    return dup

def singles(songs, artist, dic):
    di = {}
    li= []
    s= songs
    ar= truncateArtists(artist, sortedSet=False) 
    for i in range(len(ar)):
        di[ar[i]]=s[i]
    artistCount = dic
    for i in artistCount:
        if artistCount[i]==1: 
            li.append("%s - %s" %(i, di[i]))    
    return li

def getSongCountPerArtist(artists):
    ar = truncateArtists(artists, sortedSet=False)
    di={}
    for i in ar:  
        di[i]=0
    for i in ar:
        di[i]+=1    
    return di    
        
#prints out values in complex spotify spreadsheet   
def test(s, ar, al):
    li=[]
    for i in range(len(s)):
       li.append("%s\t\t\t%s\t%s" %(s[i],ar[i],al[i]))
    return li   

def prettyPrint(s, ar, al):
    li=[]
    p=[]
    for i in range(len(s)):
        li.append([s[i],ar[i],al[i]])
        print li[i]
        
    widths = [max(map(len, col)) for col in zip(*li)] 
    for row in li:
       p.append("  ".join((val.ljust(width) for val, width in zip(row, widths))))
    return p
    
def changeArtists():
    wb = openpyxl.load_workbook('playlistInfo.xlsx')
    sheet = wb.active
    s1= getSortedSet(songs=True)
    s2 = getSongs(plusArtist=False)
    a1 = getSortedSet()
    a2 = getArtists()
    di={}
    di2={}
    for i in range(len(s2)):
        di2["%s - %s" % (s2[i], a2[i][:5])]=a2[i]
        s2[i]="%s - %s" % (s2[i], a2[i][:5])
    for i in range(len(s1)):
        di["%s - %s" % (s1[i], a1[i][:5])]=a1[i]   
        s1[i]="%s - %s" % (s1[i], a1[i][:5]) 
    for i in range(len(s2)):
        if s2[i] in s1:
            sheet.cell(coordinate='B'+str(i+2), value=di[s2[i]])
    wb.save('playlistInfo.xlsx')        
    return "done"

def changeSongs():
    wb = openpyxl.load_workbook('playlistInfo.xlsx')
    sheet = wb.active
    for i in range(1, sheet.max_row+1):
        val = "%s"%sheet['A'+str(i)].value
        if val.find('- ')!=-1:  
            val = val[:val.find('- ')-1]
        if val.find('(feat')!=-1:
            val = val[:val.find('(feat')-1]
        if val.find('emaster')!=-1: 
            val = val[:val.find('emaster')-3]
        if val.lower().find('(mono')!=-1: 
            val = val[:val.lower().find('(mono')-1]
        if val.find('[')!=-1: 
            val = val[:val.lower().find('[')-1]
        sheet.cell(coordinate='A'+str(i), value=val)
    wb.save('playlistInfo.xlsx')        

def topArtists(num, dic):
    di = dic
    li = list(di)
    li = sorted(li, key = di.__getitem__)
    top = []
    count = 1
    for i in range(len(li)-1, len(li)-num-1, -1):
        top.append("%s. %s - %s" % (count, li[i], di[li[i]]))
        count+=1
    return top  

def numSongs(num, dic):
    di = dic
    li=[]
    for i in di:
        if di[i]>=num: li.append(i)
    li = sorted(li, key = di.__getitem__, reverse = True)
    for i in range(len(li)):
        li[i]="-".join([li[i], "%s"%di[li[i]]])
    return li

def getLength(str):
    if str.lower() == "artist": return len(truncateArtists(getArtists(), sortedSet=True))
    elif str.lower() == "album": return len(getAlbums(ss=True))
    elif str.lower() == "song": return len(getSongs())
    else: return "Invalid string"
    
def main():
    s=getSongs(plusArtist=False)
    ar = getArtists(ss=False)
    al= getAlbums(plusArtist=False, ss=False)
    li= prettyPrint(s, ar, al)

if __name__ == "__main__":
    main()